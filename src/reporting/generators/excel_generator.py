from datetime import datetime
from typing import Dict, Any, List
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .base import BaseReportGenerator

class ExcelReportGenerator(BaseReportGenerator):
    """
    Generates multi-sheet Excel reports with formatting.
    """
    def generate(self, summary: Dict[str, Any], output_path: str):
        self._ensure_dir(output_path)
        
        wb = openpyxl.Workbook()
        
        # 1. Summary Sheet
        self._create_summary_sheet(wb, summary)
        
        # 2. Economy Sheet
        self._create_economy_sheet(wb, summary)
        
        # 3. Military Sheet
        self._create_military_sheet(wb, summary)
        
        # 4. Tech Progression Sheet
        self._create_tech_sheet(wb, summary)
        
        # 5. Battles Sheet (if available)
        if "battles" in summary:
            self._create_battles_sheet(wb, summary)
            
        # Remove default sheet if empty
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb["Sheet"]
            
        wb.save(output_path)
        
    def _create_summary_sheet(self, wb, summary):
        ws = wb.create_sheet("Summary", 0)
        
        # Headers
        headers = ["Metric", "Value"]
        ws.append(headers)
        
        # Data
        meta = summary.get("metadata", {})
        eco = summary.get("economy", {})
        mil = summary.get("military", {})
        
        data = [
            ("Universe", summary.get("universe", "Unknown")),
            ("Turn", summary.get("turn", 0)),
            ("Timestamp", meta.get("timestamp", datetime.now().isoformat())),
            ("", ""),
            ("Total Requisition", eco.get("requisition", 0)),
            ("Income (Last Turn)", eco.get("income", 0)),
            ("Expenses (Last Turn)", eco.get("expenses", 0)),
            ("", ""),
            ("Military Power", mil.get("military_power", 0)),
            ("Total Units", mil.get("total_units", 0)),
            ("Fleets", mil.get("fleet_count", 0)),
            ("", ""),
            ("Tech Unlocked", len(summary.get("tech", {}).get("unlocked_techs", [])))
        ]
        
        for row in data:
            ws.append(row)
            
        # Formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            
        # Adjust widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20

    def _create_economy_sheet(self, wb, summary):
        ws = wb.create_sheet("Economy")
        # Assuming the summary might contain history, or we just snapshot current state
        # In a real comprehensive report, we'd iterate over history. 
        # For now, we'll list current resource breakdown if available, or just single row
        
        headers = ["Category", "Amount"]
        ws.append(headers)
        
        eco = summary.get("economy", {})
        ws.append(["Requisition", eco.get("requisition", 0)])
        ws.append(["Income", eco.get("income", 0)])
        ws.append(["Maintenance", eco.get("maintenance_cost", 0)])
        ws.append(["Construction Spend", eco.get("construction_spend", 0)])

        self._style_header(ws)

    def _create_military_sheet(self, wb, summary):
        ws = wb.create_sheet("Military")
        headers = ["Unit Type", "Count"]
        ws.append(headers)
        
        mil = summary.get("military", {})
        units = mil.get("units_by_type", {})
        
        for u_type, count in units.items():
            ws.append([u_type, count])
            
        self._style_header(ws)

    def _create_tech_sheet(self, wb, summary):
        ws = wb.create_sheet("Technology")
        headers = ["Tech Name", "Status"]
        ws.append(headers)
        
        techs = summary.get("tech", {}).get("unlocked_techs", [])
        for t in techs:
            ws.append([t, "Unlocked"])
            
        self._style_header(ws)
        ws.column_dimensions['A'].width = 40

    def _create_battles_sheet(self, wb, summary):
        ws = wb.create_sheet("Battles")
        headers = ["Location", "Winner", "Rounds", "Casualties"]
        ws.append(headers)
        
        battles = summary.get("battles", [])
        for b in battles:
            ws.append([
                b.get("location", "Unknown"),
                b.get("winner", "Unknown"),
                b.get("rounds", 0),
                b.get("casualties", 0)
            ])
            
        self._style_header(ws)

    def _style_header(self, ws):
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
